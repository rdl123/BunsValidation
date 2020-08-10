import torch
import torch.nn as nn
from fastai.vision import *
from fastai.metrics import error_rate
import shutil
import os
from openpyxl import *


def insertImage_Regular(imageFile, commentaire):
    shutil.move('../Bake_fastai_model_deploy/static/imageTesting/' + imageFile,
                '../Bake_fastai_model_deploy/Image/' + imageFile.split("@")[2])
    wb = load_workbook("../Bake_fastai_model_deploy/Train.xlsx")
    ws = wb["Feuil1"]
    imageFile = imageFile.split("@")[2]
    filename = imageFile[0:-4]
    maxrows = ws.max_row + 1
    wcell1 = ws.cell(maxrows, 1)
    wcell2 = ws.cell(maxrows, 2)
    wcell3 = ws.cell(maxrows, 3)
    wcell4 = ws.cell(maxrows, 4)
    if commentaire == "over":
        wcell1.value = filename
        wcell2.value = 1
        wcell3.value = 0
        wcell4.value = 0
        wb.save("../Bake_fastai_model_deploy/Train.xlsx")
    if commentaire == "target":
        wcell1.value = filename
        wcell2.value = 0
        wcell3.value = 1
        wcell4.value = 0
        wb.save("../Bake_fastai_model_deploy/Train.xlsx")
    if commentaire == "under":
        wcell1.value = filename
        wcell2.value = 0
        wcell3.value = 0
        wcell4.value = 1
        wb.save("../Bake_fastai_model_deploy/Train.xlsx")
    return True


def insertImage_royal(imageFile, commentaire):
    shutil.move('../Bake_fastai_model_deploy/static/imageTesting/' + imageFile,
                '../Bake_fastai_model_deploy/Royal/' + imageFile.split("@")[2])
    wb = load_workbook("../Bake_fastai_model_deploy/royal.xlsx")
    ws = wb["Feuil1"]
    imageFile = imageFile.split("@")[2]
    filename = imageFile[0:-4]
    maxrows = ws.max_row + 1
    wcell1 = ws.cell(maxrows, 1)
    wcell2 = ws.cell(maxrows, 2)
    wcell3 = ws.cell(maxrows, 3)
    wcell4 = ws.cell(maxrows, 4)
    if commentaire == "over":
        wcell1.value = filename
        wcell2.value = 1
        wcell3.value = 0
        wcell4.value = 0
        wb.save("../Bake_fastai_model_deploy/royal.xlsx")
    if commentaire == "target":
        wcell1.value = filename
        wcell2.value = 0
        wcell3.value = 1
        wcell4.value = 0
        wb.save("../Bake_fastai_model_deploy/royal.xlsx")
    if commentaire == "under":
        wcell1.value = filename
        wcell2.value = 0
        wcell3.value = 0
        wcell4.value = 1
        wb.save("../Bake_fastai_model_deploy/royal.xlsx")
    return True

def insertImage_BM(imageFile, commentaire):
    shutil.move('../Bake_fastai_model_deploy/static/imageTesting/' + imageFile,
                '../Bake_fastai_model_deploy/BM/' + imageFile.split("@")[2])
    wb = load_workbook("../Bake_fastai_model_deploy/BM.xlsx")
    ws = wb["Feuil1"]
    imageFile = imageFile.split("@")[2]
    filename = imageFile[0:-4]
    maxrows = ws.max_row + 1
    wcell1 = ws.cell(maxrows, 1)
    wcell2 = ws.cell(maxrows, 2)
    wcell3 = ws.cell(maxrows, 3)
    wcell4 = ws.cell(maxrows, 4)
    if commentaire == "over":
        wcell1.value = filename
        wcell2.value = 1
        wcell3.value= 0
        wcell4.value = 0
        wb.save("../Bake_fastai_model_deploy/BM.xlsx")
    if commentaire == "target":
        wcell1.value = filename
        wcell2.value = 0
        wcell3.value = 1
        wcell4.value = 0
        wb.save("../Bake_fastai_model_deploy/BM.xlsx")
    if commentaire == "under":
        wcell1.value = filename
        wcell2.value = 0
        wcell3.value = 0
        wcell4.value = 1
        wb.save("../Bake_fastai_model_deploy/BM.xlsx")
    return True




def retrain_model_regular():
    traindf = pd.read_excel("../Bake_fastai_model_deploy/Train.xlsx")
    traindf["image_id"] = traindf["image_id"].astype("str") + ".jpg"
    traindf["label"] = (0 * traindf.over + 1 * traindf.target + 2 * traindf.under)
    traindf.drop(columns=["over", "target", "under"], inplace=True)
    transformations = get_transforms(do_flip=True,
                                     flip_vert=True,
                                     max_lighting=0.1,
                                     max_zoom=1.05,
                                     max_warp=0.,
                                     max_rotate=15,
                                     p_affine=0.75,
                                     p_lighting=0.75
                                     )
    pathofdata = "../Bake_fastai_model_deploy/"
    data = ImageDataBunch.from_df(path=pathofdata,
                                  df=traindf,
                                  folder="Image",
                                  label_delim=None,
                                  valid_pct=0.2,
                                  seed=100,
                                  fn_col=0,
                                  label_col=1,
                                  suffix='',
                                  ds_tfms=transformations,
                                  size=480,
                                  bs=64,
                                  val_bs=32
                                  )
    data = data.normalize()
    learner = cnn_learner(data,
                          models.resnet34,
                          pretrained=True
                          , metrics=[error_rate, accuracy])

    '''
    learner.unfreeze()
    learner.lr_find(start_lr=1e-07,end_lr=0.2, num_it=100) 
    learner.freeze_to(-6)
    '''
    mingradlr1 = 0.006148692817965765
    learner.fit_one_cycle(2, slice(mingradlr1, mingradlr1 / 20))
    learner.save("trained_model", return_path=True)
    learner.load('trained_model')
    learner.model.float()
    return learner.export('model_Regular.pkl')

def retrain_model_BM():
    traindf = pd.read_excel("../Bake_fastai_model_deploy/BM.xlsx")
    traindf["image_id"] = traindf["image_id"].astype("str") + ".jpg"
    traindf["label"] = (0 * traindf.over + 1 * traindf.target + 2 * traindf.under)
    traindf.drop(columns=["over", "target", "under"], inplace=True)
    transformations = get_transforms(do_flip=True,
                                     flip_vert=True,
                                     max_lighting=0.1,
                                     max_zoom=1.05,
                                     max_warp=0.,
                                     max_rotate=15,
                                     p_affine=0.75,
                                     p_lighting=0.75
                                     )
    pathofdata = "../Bake_fastai_model_deploy/"
    data = ImageDataBunch.from_df(path=pathofdata,
                                  df=traindf,
                                  folder="BM",
                                  label_delim=None,
                                  valid_pct=0.2,
                                  seed=100,
                                  fn_col=0,
                                  label_col=1,
                                  suffix='',
                                  ds_tfms=transformations,
                                  size=480,
                                  bs=64,
                                  val_bs=32
                                  )
    data = data.normalize()
    learner = cnn_learner(data,
                          models.resnet34,
                          pretrained=True
                          , metrics=[error_rate, accuracy])

    '''
    learner.unfreeze()
    learner.lr_find(start_lr=1e-07,end_lr=0.2, num_it=100) 
    learner.freeze_to(-6)
    '''
    mingradlr1 = 0.006148692817965765
    learner.fit_one_cycle(2, slice(mingradlr1, mingradlr1 / 20))
    learner.save("trained_model_BM", return_path=True)
    learner.load('trained_model_BM')
    learner.model.float()
    return learner.export('model_bm.pkl')

def retrain_model_royal():
    traindf = pd.read_excel("../Bake_fastai_model_deploy/BM.xlsx")
    traindf["image_id"] = traindf["image_id"].astype("str") + ".jpg"
    traindf["label"] = (0 * traindf.over + 1 * traindf.target + 2 * traindf.under)
    traindf.drop(columns=["over", "target", "under"], inplace=True)
    transformations = get_transforms(do_flip=True,
                                     flip_vert=True,
                                     max_lighting=0.1,
                                     max_zoom=1.05,
                                     max_warp=0.,
                                     max_rotate=15,
                                     p_affine=0.75,
                                     p_lighting=0.75
                                     )
    pathofdata = "../Bake_fastai_model_deploy/"
    data = ImageDataBunch.from_df(path=pathofdata,
                                  df=traindf,
                                  folder="BM",
                                  label_delim=None,
                                  valid_pct=0.2,
                                  seed=100,
                                  fn_col=0,
                                  label_col=1,
                                  suffix='',
                                  ds_tfms=transformations,
                                  size=480,
                                  bs=64,
                                  val_bs=32
                                  )
    data = data.normalize()
    learner = cnn_learner(data,
                          models.resnet34,
                          pretrained=True
                          , metrics=[error_rate, accuracy])

    '''
    learner.unfreeze()
    learner.lr_find(start_lr=1e-07,end_lr=0.2, num_it=100) 
    learner.freeze_to(-6)
    '''
    mingradlr1 = 0.006148692817965765
    learner.fit_one_cycle(2, slice(mingradlr1, mingradlr1 / 20))
    learner.save("trained_model_BM", return_path=True)
    learner.load('trained_model_BM')
    learner.model.float()
    return learner.export('model_royal.pkl')